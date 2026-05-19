import os
import io
import re
import tempfile
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
import streamlit.components.v1 as components
from dotenv import load_dotenv

from parsers.pdf_parser import parse_pdf
from parsers.excel_parser import parse_excel
from parsers.normalizer import normalize_transactions
from categorizer.rules import categorize, get_all_categories
from categorizer.ai_categorizer import categorize_batch

load_dotenv()

st.set_page_config(
    page_title="Pilotto",
    page_icon="💚",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.html("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

section[data-testid="stSidebar"] {
    background: linear-gradient(160deg, #022C22 0%, #064E3B 100%);
}
section[data-testid="stSidebar"] * { color: #D1FAE5 !important; }
section[data-testid="stSidebar"] .stTextInput input {
    background: rgba(255,255,255,0.08) !important;
    border: 1px solid rgba(0,200,83,0.3) !important;
    color: #D1FAE5 !important;
    border-radius: 8px;
}

.pilotto-header {
    background: linear-gradient(135deg, #00C853 0%, #00897B 100%);
    border-radius: 16px;
    padding: 2rem 2.5rem;
    margin-bottom: 1.5rem;
}
.pilotto-header h1 {
    font-family: 'Inter', sans-serif;
    font-size: 2.2rem;
    font-weight: 800;
    margin: 0;
    letter-spacing: -0.5px;
    color: white;
}
.pilotto-header p {
    font-family: 'Inter', sans-serif;
    margin: 0.25rem 0 0;
    font-size: 1rem;
    opacity: 0.85;
    color: white;
}

.metric-card {
    background: #FFFFFF;
    border-radius: 14px;
    padding: 1.25rem 1.5rem;
    border: 1px solid #D1FAE5;
    box-shadow: 0 2px 12px rgba(0,200,83,0.07);
    border-left: 4px solid #6EE7B7;
}
.metric-positive { border-left-color: #00C853; }
.metric-negative { border-left-color: #F43F5E; }
.metric-neutral  { border-left-color: #00897B; }
.metric-label {
    font-family: 'Inter', sans-serif;
    font-size: 0.78rem;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    color: #6B7280;
    margin-bottom: 0.4rem;
}
.metric-value {
    font-family: 'Inter', sans-serif;
    font-size: 1.75rem;
    font-weight: 800;
    letter-spacing: -0.5px;
    color: #0F172A;
}
.value-positive { color: #00C853; }
.value-negative { color: #F43F5E; }

button[data-baseweb="tab"] {
    font-family: 'Inter', sans-serif !important;
    font-weight: 600 !important;
    font-size: 0.9rem !important;
}

.stButton > button {
    font-family: 'Inter', sans-serif !important;
    font-weight: 600 !important;
    border-radius: 10px !important;
    border: none !important;
    background: linear-gradient(135deg, #00C853, #00897B) !important;
    color: white !important;
    padding: 0.5rem 1.5rem !important;
    transition: opacity 0.2s;
}
.stButton > button:hover { opacity: 0.88; }

[data-testid="stFileUploader"] {
    border: 2px dashed #6EE7B7 !important;
    border-radius: 14px !important;
    background: #F0FDF4 !important;
}
</style>
""")


# ── Session state initialisation ────────────────────────────────────────────

DEFAULT_CATEGORY_TYPES = {
    "Alimentação": "Saída", "Transporte": "Saída", "Moradia": "Saída",
    "Saúde": "Saída", "Lazer": "Saída", "Compras": "Saída",
    "Educação": "Saída", "Serviços": "Saída", "Outros": "Saída",
    "Receita": "Entrada",
}


def _init_state():
    defaults = {
        "transactions_df": pd.DataFrame(),
        "uploaded_files_names": [],
        "api_key": os.getenv("ANTHROPIC_API_KEY", ""),
        "categorization_done": False,
        "wizard_step": 1,
        "historico": [],
        "custom_categories": [],
        # {name: "Entrada" | "Saída"}
        "category_types": dict(DEFAULT_CATEGORY_TYPES),
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


_init_state()


def _get_categories() -> list[str]:
    base = get_all_categories()
    custom = st.session_state.get("custom_categories", [])
    merged = base + [c for c in custom if c not in base]
    return sorted(merged)


def _categories_by_type(tipo: str) -> list[str]:
    types = st.session_state.get("category_types", {})
    all_cats = _get_categories()
    filtered = [c for c in all_cats if types.get(c, "Saída") == tipo]
    return filtered if filtered else all_cats


def _get_cat_type(name: str) -> str:
    return st.session_state.get("category_types", {}).get(name, "Saída")


# ── Sidebar ──────────────────────────────────────────────────────────────────

with st.sidebar:
    st.title("⚙️ Configurações")
    st.divider()

    api_key_input = st.text_input(
        "Chave API Anthropic (opcional)",
        value=st.session_state.api_key,
        type="password",
        help="Usada para categorização com IA. Se vazia, apenas regras locais são aplicadas.",
    )
    if api_key_input != st.session_state.api_key:
        st.session_state.api_key = api_key_input

    if st.session_state.api_key:
        st.success("✅ API key configurada")
    else:
        st.info("ℹ️ Sem API key — usando apenas regras locais")

    st.divider()

    if st.session_state.uploaded_files_names:
        st.subheader("📂 Arquivos carregados")
        for name in st.session_state.uploaded_files_names:
            st.text(f"• {name}")

        if st.button("🗑️ Limpar todos os dados", use_container_width=True):
            st.session_state.transactions_df = pd.DataFrame()
            st.session_state.uploaded_files_names = []
            st.session_state.categorization_done = False
            st.session_state.wizard_step = 1
            st.rerun()
    else:
        st.caption("Nenhum arquivo carregado ainda.")

    st.divider()
    st.caption("Pilotto v1.0 — Seu dinheiro no pilotto automático")


# ── Helpers ──────────────────────────────────────────────────────────────────

def _process_file(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name
    suffix = name.rsplit(".", 1)[-1].lower()

    with tempfile.NamedTemporaryFile(delete=False, suffix=f".{suffix}") as tmp:
        tmp.write(uploaded_file.getvalue())
        tmp_path = tmp.name

    try:
        if suffix == "pdf":
            raw = parse_pdf(tmp_path)
        elif suffix in ("xlsx", "xls", "csv"):
            raw = parse_excel(tmp_path)
        else:
            raise ValueError(f"Formato não suportado: .{suffix}")
    finally:
        os.unlink(tmp_path)

    return normalize_transactions(raw, source_name=name)


def _apply_categorization(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    categories = _get_categories()
    result_cats = []

    for _, row in df.iterrows():
        cat = categorize(row["description"])
        result_cats.append(cat if cat else "__UNCATEGORIZED__")

    uncategorized_idx = [i for i, c in enumerate(result_cats) if c == "__UNCATEGORIZED__"]

    if uncategorized_idx and st.session_state.api_key:
        uncategorized_txns = [
            df.iloc[i].to_dict() for i in uncategorized_idx
        ]
        with st.spinner(f"Categorizando {len(uncategorized_txns)} transações com IA..."):
            ai_cats = categorize_batch(
                uncategorized_txns, categories, api_key=st.session_state.api_key
            )
        for list_pos, df_idx in enumerate(uncategorized_idx):
            result_cats[df_idx] = ai_cats[list_pos]
    else:
        for i in uncategorized_idx:
            result_cats[i] = "Outros"

    df = df.copy()
    df["categoria"] = result_cats
    return df


def _extract_favorecido(description: str) -> str:
    """Extrai o nome do pagador ou recebedor a partir da descrição da transação."""
    if not description:
        return ""
    desc = description.strip()

    patterns = [
        # PIX / TED / DOC recebido de NOME
        r"(?:pix|ted|doc)\s+(?:recebido|recebida)\s+de\s+(.+?)(?:\s*[-–|/]|\s*cpf\b|\s*\d{3}\.|\s*$)",
        # PIX / TED / DOC enviado / para NOME
        r"(?:pix|ted|doc)\s+(?:enviado|enviada)?\s*(?:para\s+)?(.+?)(?:\s*[-–|/]|\s*cpf\b|\s*\d{3}\.|\s*$)",
        # Transferência recebida de / para NOME
        r"transfer[eê]ncia\s+recebida\s+de\s+(.+?)(?:\s*[-–|/]|\s*\d|$)",
        r"transfer[eê]ncia\s+(?:para|enviada\s+para)\s+(.+?)(?:\s*[-–|/]|\s*\d|$)",
        # Formatos "TIPO - NOME" (Bradesco, Itaú)
        r"^(?:pix|ted|doc|pagamento)\s*[-–]\s*(.+?)(?:\s*[-–]|\s*$)",
        # Compra débito / crédito NOME
        r"^compra\s+(?:cart[aã]o\s+)?(?:d[eé]bito|cr[eé]dito)?\s*[-–]?\s*(.+?)(?:\s*[-–]|\s*$)",
        # Pagamento para / de NOME
        r"^pagamento\s+(?:a\s+|para\s+|de\s+)?(.+?)(?:\s*[-–|/]|\s*\d|$)",
        # Débito automático NOME
        r"^d[eé]bito\s+(?:autom[aá]tico\s+)?(.+?)(?:\s*[-–]|\s*$)",
    ]

    for pattern in patterns:
        m = re.search(pattern, desc, re.IGNORECASE)
        if m:
            name = m.group(1).strip()
            # Remove CPF/CNPJ inline
            name = re.sub(r"\s*cpf.*$", "", name, flags=re.IGNORECASE)
            name = re.sub(r"\s*\d{3}\.\d{3}\.\d{3}-\d{2}.*$", "", name)
            name = re.sub(r"\s*\d{2}\.\d{3}\.\d{3}.*$", "", name)
            name = name.strip("-– /").strip()
            if len(name) > 2:
                return name.title()

    # Fallback: segundo segmento após " - "
    if " - " in desc:
        parts = desc.split(" - ")
        for part in parts[1:]:
            part = part.strip()
            if len(part) > 3 and not re.match(r"^\d", part):
                return part.title()

    return ""


def _format_currency(value: float) -> str:
    return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _color_value(value: float) -> str:
    formatted = _format_currency(value)
    cls = "value-positive" if value >= 0 else "value-negative"
    return f'<span class="{cls}">{formatted}</span>'


# ── Tab: Upload (wizard) ──────────────────────────────────────────────────────

def _wizard_progress(step: int):
    steps = ["📤 Upload", "📥 Entradas", "📤 Saídas"]
    cols = st.columns(len(steps))
    for i, (col, label) in enumerate(zip(cols, steps), start=1):
        if i < step:
            col.success(f"✓ {label}")
        elif i == step:
            col.info(f"**→ {label}**")
        else:
            col.markdown(f"<span style='color:#9CA3AF'>{label}</span>", unsafe_allow_html=True)


def render_upload_tab():
    step = st.session_state.wizard_step

    _wizard_progress(step)
    st.divider()

    # ── Passo 1: Upload ──
    if step == 1:
        st.subheader("Suba seus extratos")
        st.caption("Extrato da conta corrente e fatura do cartão — pode enviar os dois juntos.")

        uploaded_files = st.file_uploader(
            "Arraste os arquivos ou clique para selecionar",
            type=["pdf", "xlsx", "xls", "csv"],
            accept_multiple_files=True,
            label_visibility="collapsed",
        )

        if uploaded_files:
            new_files = [f for f in uploaded_files if f.name not in st.session_state.uploaded_files_names]

            if new_files:
                progress_bar = st.progress(0, text="Lendo arquivos...")
                new_dfs, errors = [], []

                for idx, uploaded_file in enumerate(new_files):
                    progress_bar.progress((idx + 1) / len(new_files), text=f"Lendo {uploaded_file.name}...")
                    try:
                        new_dfs.append(_process_file(uploaded_file))
                        st.session_state.uploaded_files_names.append(uploaded_file.name)
                    except Exception as e:
                        errors.append((uploaded_file.name, str(e)))

                progress_bar.empty()
                for name, err in errors:
                    st.error(f"**{name}**: {err}")

                if new_dfs:
                    combined = pd.concat(new_dfs, ignore_index=True)
                    if not st.session_state.transactions_df.empty:
                        combined = pd.concat(
                            [st.session_state.transactions_df, combined], ignore_index=True
                        ).drop_duplicates(subset=["date", "description", "value"]).reset_index(drop=True)

                    # Extrai favorecido para linhas ainda sem ele
                    if "favorecido" not in combined.columns:
                        combined["favorecido"] = ""
                    mask_fav = combined["favorecido"].isna() | (combined["favorecido"] == "")
                    combined.loc[mask_fav, "favorecido"] = combined.loc[mask_fav, "description"].apply(_extract_favorecido)

                    with st.spinner("Categorizando automaticamente..."):
                        if "categoria" not in combined.columns:
                            combined["categoria"] = None
                        needs_cat = combined["categoria"].isna() | (combined["categoria"] == "")
                        if needs_cat.any():
                            sub_cat = _apply_categorization(combined[needs_cat].copy())
                            combined.loc[needs_cat, "categoria"] = sub_cat["categoria"].values

                    st.session_state.transactions_df = combined
                    st.session_state.categorization_done = True

        df = st.session_state.transactions_df
        if not df.empty:
            n_in = (df["value"] > 0).sum()
            n_out = (df["value"] <= 0).sum()
            st.success(f"✅ {len(df)} transações reconhecidas — {n_in} entradas, {n_out} saídas.")

            with st.expander("🔍 Inspecionar transações detectadas"):
                debug_df = df.copy()
                debug_df["data"] = debug_df["date"].dt.strftime("%d/%m/%Y")
                debug_df["valor"] = debug_df["value"].apply(lambda v: f"+{_format_currency(v)}" if v > 0 else _format_currency(v))
                st.dataframe(
                    debug_df[["data", "description", "valor", "type", "source"]].rename(
                        columns={"description": "descrição", "type": "tipo detectado", "source": "arquivo"}
                    ),
                    use_container_width=True,
                    hide_index=True,
                    height=300,
                )
                st.caption("tipo detectado: 'credit' = entrada  |  'debit' = saída")

            if st.button("Classificar Transações →", use_container_width=True):
                st.session_state.wizard_step = 2
                st.rerun()

    # ── Passo 2: Revisar Entradas ──
    elif step == 2:
        df = st.session_state.transactions_df
        mask_in = df["value"] > 0
        entradas = df[mask_in].copy()

        if entradas.empty:
            st.info("Nenhuma entrada encontrada nos arquivos. Avançando para as saídas.")
            st.session_state.wizard_step = 3
            st.rerun()
        else:
            st.subheader(f"Revisar Entradas ({len(entradas)} transações)")
            st.caption("Confirme ou corrija a categoria de cada receita.")

            categories = _categories_by_type("Entrada")
            edit_df = entradas[["date", "description", "value", "categoria"]].copy()
            edit_df["favorecido"] = entradas.get("favorecido", pd.Series("", index=entradas.index)).values
            edit_df["data"] = edit_df["date"].dt.strftime("%d/%m/%Y")
            edit_df["valor"] = edit_df["value"].abs()
            edit_df = edit_df[["data", "description", "favorecido", "valor", "categoria"]].rename(columns={"description": "descrição"})
            edit_df["_idx"] = entradas.index

            edited = st.data_editor(
                edit_df,
                column_config={
                    "data": st.column_config.TextColumn("Data", disabled=True),
                    "descrição": st.column_config.TextColumn("Descrição", disabled=True),
                    "favorecido": st.column_config.TextColumn("Origem", disabled=True),
                    "valor": st.column_config.NumberColumn("Valor (R$)", format="R$ %.2f"),
                    "categoria": st.column_config.TextColumn("Categoria", help="Digite ou edite a categoria — novas serão salvas automaticamente"),
                    "_idx": None,
                },
                use_container_width=True,
                hide_index=True,
                num_rows="dynamic",
                height=420,
            )

            if st.button("Confirmar Entradas →", use_container_width=True):
                kept = edited["_idx"].dropna().astype(int).tolist()
                deleted = [i for i in entradas.index if i not in kept]
                df = df.drop(index=deleted)
                existing = _get_categories()
                for _, row in edited.iterrows():
                    if pd.notna(row.get("_idx")):
                        idx = int(row["_idx"])
                        cat = str(row["categoria"]).strip() if pd.notna(row.get("categoria")) else "Receita"
                        if cat and cat not in existing:
                            st.session_state.custom_categories.append(cat)
                            st.session_state.category_types[cat] = "Entrada"
                            existing.append(cat)
                        df.at[idx, "categoria"] = cat
                        df.at[idx, "value"] = abs(row["valor"])
                st.session_state.transactions_df = df.reset_index(drop=True)
                st.session_state.wizard_step = 3
                st.rerun()

    # ── Passo 3: Revisar Saídas ──
    elif step == 3:
        df = st.session_state.transactions_df
        mask_out = df["value"] < 0
        saidas = df[mask_out].copy()

        if saidas.empty:
            st.info("Nenhuma saída encontrada nos arquivos.")
            if st.button("Ver Dashboard →", use_container_width=True):
                st.rerun()
        else:
            st.subheader(f"Revisar Saídas ({len(saidas)} transações)")
            st.caption("Confirme ou corrija a categoria de cada gasto.")

            categories = _categories_by_type("Saída")
            edit_df = saidas[["date", "description", "value", "categoria"]].copy()
            edit_df["favorecido"] = saidas.get("favorecido", pd.Series("", index=saidas.index)).values
            edit_df["data"] = edit_df["date"].dt.strftime("%d/%m/%Y")
            edit_df["valor"] = edit_df["value"].abs()
            edit_df = edit_df[["data", "description", "favorecido", "valor", "categoria"]].rename(columns={"description": "descrição"})
            edit_df["_idx"] = saidas.index

            edited = st.data_editor(
                edit_df,
                column_config={
                    "data": st.column_config.TextColumn("Data", disabled=True),
                    "descrição": st.column_config.TextColumn("Descrição", disabled=True),
                    "favorecido": st.column_config.TextColumn("Destino", disabled=True),
                    "valor": st.column_config.NumberColumn("Valor (R$)", format="R$ %.2f"),
                    "categoria": st.column_config.TextColumn("Categoria", help="Digite ou edite a categoria — novas serão salvas automaticamente"),
                    "_idx": None,
                },
                use_container_width=True,
                hide_index=True,
                num_rows="dynamic",
                height=420,
            )

            if st.button("Finalizar e Ver Dashboard →", use_container_width=True):
                kept = edited["_idx"].dropna().astype(int).tolist()
                deleted = [i for i in saidas.index if i not in kept]
                df = df.drop(index=deleted)
                existing = _get_categories()
                for _, row in edited.iterrows():
                    if pd.notna(row.get("_idx")):
                        idx = int(row["_idx"])
                        cat = str(row["categoria"]).strip() if pd.notna(row.get("categoria")) else "Outros"
                        if cat and cat not in existing:
                            st.session_state.custom_categories.append(cat)
                            st.session_state.category_types[cat] = "Saída"
                            existing.append(cat)
                        df.at[idx, "categoria"] = cat
                        df.at[idx, "value"] = -abs(row["valor"])
                st.session_state.transactions_df = df.reset_index(drop=True)
                st.rerun()


# ── Tab: Dashboard ────────────────────────────────────────────────────────────

def render_dashboard_tab():
    st.header("📊 Dashboard")

    df = st.session_state.transactions_df
    if df.empty:
        st.info("Carregue extratos na aba **Upload** para ver o dashboard.")
        return

    df = df.copy()
    df["month"] = df["date"].dt.to_period("M")

    months = sorted(df["month"].unique())
    month_options = ["Todos os meses"] + [str(m) for m in months]

    selected_month_str = st.selectbox(
        "Período",
        options=month_options,
        index=0,
    )

    if selected_month_str != "Todos os meses":
        selected_period = pd.Period(selected_month_str, freq="M")
        filtered = df[df["month"] == selected_period]
    else:
        filtered = df

    receitas = filtered[filtered["value"] > 0]["value"].sum()
    despesas = filtered[filtered["value"] < 0]["value"].sum()
    saldo = receitas + despesas

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric(
            label="💚 Total Receitas",
            value=_format_currency(receitas),
        )
    with col2:
        st.metric(
            label="🔴 Total Despesas",
            value=_format_currency(abs(despesas)),
        )
    with col3:
        delta_color = "normal" if saldo >= 0 else "inverse"
        st.metric(
            label="Saldo do Mês",
            value=_format_currency(saldo),
            delta=f"{'positivo' if saldo >= 0 else 'negativo'}",
            delta_color=delta_color,
        )

    st.divider()

    expenses_df = filtered[filtered["value"] < 0].copy()
    expenses_df["abs_value"] = expenses_df["value"].abs()

    if expenses_df.empty:
        st.info("Nenhuma despesa no período selecionado.")
    else:
        cat_summary = (
            expenses_df.groupby("categoria")["abs_value"]
            .sum()
            .reset_index()
            .sort_values("abs_value", ascending=False)
        )

        col_pie, col_bar = st.columns(2)

        with col_pie:
            st.subheader("Despesas por Categoria")
            fig_pie = px.pie(
                cat_summary,
                names="categoria",
                values="abs_value",
                hole=0.4,
                color_discrete_sequence=px.colors.qualitative.Set3,
            )
            fig_pie.update_traces(
                textposition="inside",
                textinfo="percent+label",
                hovertemplate="<b>%{label}</b><br>R$ %{value:,.2f}<br>%{percent}<extra></extra>",
            )
            fig_pie.update_layout(
                showlegend=True,
                legend=dict(orientation="v", yanchor="middle", y=0.5),
                margin=dict(t=10, b=10, l=10, r=10),
                height=380,
            )
            st.plotly_chart(fig_pie, use_container_width=True)

        with col_bar:
            st.subheader("Valor por Categoria")
            fig_bar = px.bar(
                cat_summary.sort_values("abs_value"),
                x="abs_value",
                y="categoria",
                orientation="h",
                color="abs_value",
                color_continuous_scale="Reds",
                labels={"abs_value": "Valor (R$)", "categoria": "Categoria"},
            )
            fig_bar.update_traces(
                hovertemplate="<b>%{y}</b><br>R$ %{x:,.2f}<extra></extra>"
            )
            fig_bar.update_layout(
                showlegend=False,
                coloraxis_showscale=False,
                margin=dict(t=10, b=10, l=10, r=10),
                height=380,
                xaxis_title="Valor (R$)",
                yaxis_title="",
            )
            st.plotly_chart(fig_bar, use_container_width=True)

    if len(months) > 1:
        st.divider()
        st.subheader("Evolução Mensal")

        monthly = (
            df.groupby(["month", df["value"].apply(lambda v: "Receita" if v > 0 else "Despesa").rename("tipo")])
            ["value"]
            .sum()
            .abs()
            .reset_index()
        )
        monthly.columns = ["month", "tipo", "valor"]
        monthly["month_str"] = monthly["month"].astype(str)

        fig_trend = px.bar(
            monthly,
            x="month_str",
            y="valor",
            color="tipo",
            barmode="group",
            color_discrete_map={"Receita": "#28a745", "Despesa": "#dc3545"},
            labels={"month_str": "Mês", "valor": "Valor (R$)", "tipo": "Tipo"},
        )
        fig_trend.update_traces(
            hovertemplate="<b>%{x}</b><br>R$ %{y:,.2f}<extra></extra>"
        )
        fig_trend.update_layout(
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
            margin=dict(t=30, b=10, l=10, r=10),
            height=350,
        )
        st.plotly_chart(fig_trend, use_container_width=True)


# ── Tab: Categorias ───────────────────────────────────────────────────────────

def render_categorias_tab():
    st.header("🏷️ Categorias")
    st.caption("Crie e organize suas categorias. O tipo define em qual etapa elas aparecem.")

    types = st.session_state.category_types
    base_cats = get_all_categories()
    custom_cats = st.session_state.custom_categories

    # ── Criar nova categoria ──
    with st.expander("➕ Nova categoria", expanded=True):
        c1, c2, c3 = st.columns([3, 2, 1])
        with c1:
            novo_nome = st.text_input("Nome", key="cat_new_name", placeholder="Ex: Academia, Investimentos...")
        with c2:
            novo_tipo = st.selectbox("Tipo", ["Saída", "Entrada"], key="cat_new_type")
        with c3:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("Adicionar", key="cat_add_btn", use_container_width=True):
                nome = novo_nome.strip()
                if nome and nome not in _get_categories():
                    st.session_state.custom_categories.append(nome)
                    st.session_state.category_types[nome] = novo_tipo
                    st.rerun()
                elif nome in _get_categories():
                    st.warning(f'"{nome}" já existe.')

    st.divider()

    # ── Listar todas as categorias ──
    all_cats = _get_categories()
    entradas = [c for c in all_cats if types.get(c, "Saída") == "Entrada"]
    saidas   = [c for c in all_cats if types.get(c, "Saída") == "Saída"]

    for label, cats, tipo_val in [("📥 Entradas", entradas, "Entrada"), ("📤 Saídas", saidas, "Saída")]:
        st.subheader(label)
        for cat in cats:
            is_custom = cat in custom_cats
            c1, c2, c3 = st.columns([3, 2, 1])
            with c1:
                if is_custom:
                    novo = st.text_input("", value=cat, key=f"cat_name_{cat}", label_visibility="collapsed")
                else:
                    st.markdown(f"**{cat}**")
                    novo = cat
            with c2:
                tipo_atual = types.get(cat, "Saída")
                novo_tipo = st.selectbox("", ["Saída", "Entrada"], index=0 if tipo_atual == "Saída" else 1,
                                         key=f"cat_type_{cat}", label_visibility="collapsed")
                if novo_tipo != tipo_atual:
                    st.session_state.category_types[cat] = novo_tipo
                    st.rerun()
            with c3:
                if is_custom:
                    if st.button("✕", key=f"cat_del_{cat}", help="Remover"):
                        st.session_state.custom_categories.remove(cat)
                        st.session_state.category_types.pop(cat, None)
                        st.rerun()
                    if novo != cat and novo.strip():
                        st.session_state.custom_categories[st.session_state.custom_categories.index(cat)] = novo.strip()
                        st.session_state.category_types[novo.strip()] = st.session_state.category_types.pop(cat)
                        st.rerun()
                else:
                    st.markdown("<span style='color:#9CA3AF;font-size:0.75rem'>padrão</span>", unsafe_allow_html=True)
        st.markdown("")


# ── Tab: Início ───────────────────────────────────────────────────────────────

def render_inicio_tab():
    df = st.session_state.transactions_df
    historico = st.session_state.historico

    if df.empty and not historico:
        st.markdown("### Bem-vindo ao Pilotto")
        st.markdown("Seu assistente de finanças pessoais. Comece subindo o extrato do mês.")
        if st.button("➕ Subir extrato agora", use_container_width=True):
            st.session_state["_goto_upload"] = True
            st.rerun()
        return

    if not df.empty:
        receitas = df[df["value"] > 0]["value"].sum()
        despesas = df[df["value"] <= 0]["value"].sum()
        saldo = receitas + despesas
        saldo_class = "metric-positive" if saldo >= 0 else "metric-negative"
        saldo_color = "#00C853" if saldo >= 0 else "#F43F5E"

        c1, c2, c3 = st.columns(3)
        with c1:
            st.html(f"""<div class="metric-card metric-positive">
                <div class="metric-label">Receitas do Mês</div>
                <div class="metric-value value-positive">{_format_currency(receitas)}</div>
            </div>""")
        with c2:
            st.html(f"""<div class="metric-card metric-negative">
                <div class="metric-label">Despesas do Mês</div>
                <div class="metric-value value-negative">{_format_currency(abs(despesas))}</div>
            </div>""")
        with c3:
            st.html(f"""<div class="metric-card {saldo_class}">
                <div class="metric-label">Saldo do Mês</div>
                <div class="metric-value" style="color:{saldo_color}">{_format_currency(saldo)}</div>
            </div>""")

        st.divider()

    if historico:
        ultimo = historico[-1]
        saldo_ult = ultimo["receitas"] - ultimo["despesas"]
        saldo_color = "#00C853" if saldo_ult >= 0 else "#F43F5E"
        st.markdown(f"**Último mês fechado:** {ultimo['mes']}")
        c1, c2, c3 = st.columns(3)
        with c1:
            st.metric("Receitas", _format_currency(ultimo["receitas"]))
        with c2:
            st.metric("Despesas", _format_currency(ultimo["despesas"]))
        with c3:
            st.metric("Saldo", _format_currency(saldo_ult), delta=None)


# ── Tab: Histórico ─────────────────────────────────────────────────────────────

def render_historico_tab():
    import json

    st.header("📅 Histórico Mensal")

    col_imp, col_exp = st.columns([1, 1])

    with col_imp:
        hist_file = st.file_uploader("Carregar histórico salvo (.json)", type=["json"], label_visibility="collapsed", key="hist_upload")
        if hist_file:
            try:
                loaded = json.loads(hist_file.read())
                st.session_state.historico = loaded
                st.success(f"✅ {len(loaded)} meses carregados.")
                st.rerun()
            except Exception:
                st.error("Arquivo inválido.")

    historico = st.session_state.historico

    df = st.session_state.transactions_df
    if not df.empty:
        st.divider()
        st.subheader("Fechar mês atual")
        meses_disponiveis = sorted(df["date"].dt.to_period("M").unique(), reverse=True)
        mes_opts = [str(m) for m in meses_disponiveis]
        mes_sel = st.selectbox("Selecione o mês para fechar", mes_opts)

        if st.button(f"Fechar {mes_sel} e salvar no histórico", use_container_width=True):
            period = pd.Period(mes_sel, "M")
            df_mes = df[df["date"].dt.to_period("M") == period]
            receitas = float(df_mes[df_mes["value"] > 0]["value"].sum())
            despesas = float(abs(df_mes[df_mes["value"] <= 0]["value"].sum()))
            top_cats = (
                df_mes[df_mes["value"] <= 0]
                .groupby("categoria")["value"]
                .sum()
                .abs()
                .sort_values(ascending=False)
                .head(3)
                .to_dict()
            )
            entry = {
                "mes": mes_sel,
                "receitas": receitas,
                "despesas": despesas,
                "saldo": receitas - despesas,
                "top_categorias": {k: round(v, 2) for k, v in top_cats.items()},
            }
            existing = [h for h in historico if h["mes"] != mes_sel]
            existing.append(entry)
            existing.sort(key=lambda x: x["mes"])
            st.session_state.historico = existing
            st.success(f"✅ {mes_sel} fechado e salvo no histórico!")
            st.rerun()

    st.divider()

    if not historico:
        st.info("Nenhum mês fechado ainda. Suba um extrato e clique em **Fechar mês** acima.")
        return

    with col_exp:
        hist_json = json.dumps(historico, ensure_ascii=False, indent=2)
        st.download_button(
            "⬇️ Exportar histórico (.json)",
            data=hist_json,
            file_name="pilotto_historico.json",
            mime="application/json",
            use_container_width=True,
        )

    for entry in reversed(historico):
        saldo = entry["saldo"]
        saldo_color = "#00C853" if saldo >= 0 else "#F43F5E"
        top_str = "  •  ".join(
            f"{cat}: {_format_currency(v)}" for cat, v in entry.get("top_categorias", {}).items()
        )
        st.html(f"""
        <div class="metric-card" style="margin-bottom:0.75rem">
            <div style="display:flex;justify-content:space-between;align-items:center">
                <div style="font-family:Inter,sans-serif;font-weight:700;font-size:1.05rem">{entry["mes"]}</div>
                <div style="font-family:Inter,sans-serif;font-weight:800;font-size:1.1rem;color:{saldo_color}">{_format_currency(saldo)}</div>
            </div>
            <div style="display:flex;gap:2rem;margin-top:0.4rem;font-family:Inter,sans-serif;font-size:0.85rem;color:#6B7280">
                <span>🟢 Receitas: {_format_currency(entry["receitas"])}</span>
                <span>🔴 Despesas: {_format_currency(entry["despesas"])}</span>
            </div>
            {"<div style='margin-top:0.35rem;font-family:Inter,sans-serif;font-size:0.8rem;color:#9CA3AF'>" + top_str + "</div>" if top_str else ""}
        </div>
        """)

    if len(historico) >= 2:
        st.divider()
        st.subheader("Evolução")
        hist_df = pd.DataFrame(historico)
        fig = px.bar(
            hist_df.melt(id_vars="mes", value_vars=["receitas", "despesas"], var_name="tipo", value_name="valor"),
            x="mes", y="valor", color="tipo", barmode="group",
            color_discrete_map={"receitas": "#00C853", "despesas": "#F43F5E"},
            labels={"mes": "Mês", "valor": "R$", "tipo": ""},
        )
        fig.update_layout(plot_bgcolor="white", paper_bgcolor="white", font_family="Inter")
        st.plotly_chart(fig, use_container_width=True)


# ── Tab: Transações ───────────────────────────────────────────────────────────

def render_transactions_tab():
    st.header("📋 Transações")

    df = st.session_state.transactions_df
    if df.empty:
        st.info("Carregue extratos na aba **Upload** para ver as transações.")
        return

    categories = _get_categories()
    df = df.copy()

    with st.expander("➕ Gerenciar categorias"):
        col_new, col_btn = st.columns([3, 1])
        with col_new:
            nova = st.text_input("Nome da nova categoria", key="nova_cat_input", label_visibility="collapsed", placeholder="Ex: Academia, Assinaturas...")
        with col_btn:
            if st.button("Adicionar", key="btn_add_cat"):
                nome = nova.strip()
                if nome and nome not in st.session_state.custom_categories and nome not in get_all_categories():
                    st.session_state.custom_categories.append(nome)
                    st.rerun()
        custom = st.session_state.get("custom_categories", [])
        if custom:
            st.caption("Categorias criadas por você:")
            for cat in list(custom):
                c1, c2 = st.columns([5, 1])
                c1.markdown(f"• {cat}")
                if c2.button("✕", key=f"del_cat_{cat}"):
                    st.session_state.custom_categories.remove(cat)
                    st.rerun()

    st.subheader("Filtros")
    filter_col1, filter_col2, filter_col3 = st.columns(3)

    with filter_col1:
        all_cats = ["Todas"] + sorted(df["categoria"].dropna().unique().tolist())
        selected_cat = st.selectbox("Categoria", options=all_cats)

    with filter_col2:
        type_options = {"Todos": None, "Débitos (despesas)": "debit", "Créditos (receitas)": "credit"}
        selected_type_label = st.selectbox("Tipo", options=list(type_options.keys()))
        selected_type = type_options[selected_type_label]

    with filter_col3:
        min_date = df["date"].min().date()
        max_date = df["date"].max().date()
        date_range = st.date_input(
            "Período",
            value=(min_date, max_date),
            min_value=min_date,
            max_value=max_date,
        )

    filtered = df.copy()

    if selected_cat != "Todas":
        filtered = filtered[filtered["categoria"] == selected_cat]

    if selected_type:
        filtered = filtered[filtered["type"] == selected_type]

    if isinstance(date_range, (list, tuple)) and len(date_range) == 2:
        start_date, end_date = date_range
        filtered = filtered[
            (filtered["date"].dt.date >= start_date) &
            (filtered["date"].dt.date <= end_date)
        ]

    filtered = filtered.sort_values("date", ascending=False).reset_index(drop=True)

    st.caption(f"{len(filtered)} transação(ões) encontrada(s)")

    display_df = filtered.copy()
    display_df["Data"] = display_df["date"].dt.strftime("%d/%m/%Y")
    display_df["Descrição"] = display_df["description"]
    display_df["Origem / Destino"] = display_df.get("favorecido", pd.Series("", index=display_df.index)).fillna("")
    display_df["Valor (R$)"] = display_df["value"]
    display_df["Categoria"] = display_df["categoria"]
    display_df["_idx"] = filtered.index

    edited = st.data_editor(
        display_df[["Data", "Descrição", "Origem / Destino", "Valor (R$)", "Categoria", "_idx"]],
        use_container_width=True,
        hide_index=True,
        height=500,
        num_rows="dynamic",
        column_config={
            "Data": st.column_config.TextColumn("Data", width="small", disabled=True),
            "Descrição": st.column_config.TextColumn("Descrição", width="large", disabled=True),
            "Origem / Destino": st.column_config.TextColumn("Origem / Destino", width="medium", disabled=True),
            "Valor (R$)": st.column_config.NumberColumn(
                "Valor (R$)",
                format="R$ %.2f",
                width="small",
                help="Positivo = entrada, negativo = saída",
            ),
            "Categoria": st.column_config.TextColumn("Categoria", width="medium", help="Digite ou edite — novas categorias são salvas automaticamente"),
            "_idx": None,
        },
    )

    if st.button("💾 Salvar alterações"):
        kept_idxs = edited["_idx"].dropna().astype(int).tolist()
        deleted_idxs = [i for i in filtered.index if i not in kept_idxs]
        main_df = st.session_state.transactions_df.drop(index=deleted_idxs)
        existing = _get_categories()
        for _, row in edited.iterrows():
            if pd.notna(row.get("_idx")):
                idx = int(row["_idx"])
                cat = str(row["Categoria"]).strip() if pd.notna(row.get("Categoria")) else "Outros"
                if cat and cat not in existing:
                    st.session_state.custom_categories.append(cat)
                    st.session_state.category_types[cat] = "Saída"
                    existing.append(cat)
                main_df.at[idx, "categoria"] = cat
                main_df.at[idx, "value"] = row["Valor (R$)"]
        st.session_state.transactions_df = main_df.reset_index(drop=True)
        st.success(f"✅ Salvo! {len(deleted_idxs)} excluída(s)." if deleted_idxs else "✅ Alterações salvas!")
        st.rerun()

    st.divider()

    if st.button("📥 Exportar para Excel", use_container_width=False):
        export_df = filtered.copy()
        export_df["data"] = export_df["date"].dt.strftime("%d/%m/%Y")
        if "favorecido" not in export_df.columns:
            export_df["favorecido"] = ""
        export_df = export_df.rename(columns={
            "description": "descrição",
            "value": "valor",
            "type": "tipo",
            "source": "arquivo",
            "categoria": "categoria",
            "favorecido": "origem / destino",
        })
        export_df["tipo"] = export_df["tipo"].map({"debit": "Débito", "credit": "Crédito"})
        export_df = export_df[["data", "descrição", "origem / destino", "valor", "tipo", "categoria", "arquivo"]]

        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            export_df.to_excel(writer, index=False, sheet_name="Transações")

            workbook = writer.book
            worksheet = writer.sheets["Transações"]

            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            col_widths = {"data": 14, "descrição": 50, "origem / destino": 28, "valor": 16, "tipo": 12, "categoria": 18, "arquivo": 30}
            for col_idx, col_name in enumerate(export_df.columns, 1):
                width = col_widths.get(col_name, 20)
                worksheet.column_dimensions[get_column_letter(col_idx)].width = width

            from openpyxl.styles import numbers as xl_numbers
            value_col_idx = list(export_df.columns).index("valor") + 1
            for row in worksheet.iter_rows(min_row=2, min_col=value_col_idx, max_col=value_col_idx):
                for cell in row:
                    cell.number_format = '#,##0.00'

        buffer.seek(0)
        st.download_button(
            label="⬇️ Baixar arquivo Excel",
            data=buffer,
            file_name="transacoes_financeiras.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


# ── Main layout ───────────────────────────────────────────────────────────────

st.html("""
<div class="pilotto-header">
    <h1>Pilotto</h1>
    <p>Seu dinheiro no pilotto automático — analise extratos, categorize gastos e entenda seu mês.</p>
</div>
""")

tab_inicio, tab_upload, tab_dashboard, tab_transactions, tab_historico, tab_categorias = st.tabs(
    ["🏠 Início", "📤 Upload", "📊 Dashboard", "📋 Transações", "📅 Histórico", "🏷️ Categorias"]
)

with tab_inicio:
    render_inicio_tab()

with tab_upload:
    render_upload_tab()

with tab_dashboard:
    render_dashboard_tab()

with tab_transactions:
    render_transactions_tab()

with tab_historico:
    render_historico_tab()

with tab_categorias:
    render_categorias_tab()

if st.session_state.get("_goto_upload"):
    st.session_state["_goto_upload"] = False
    components.html("""
    <script>
        setTimeout(function() {
            var tabs = window.parent.document.querySelectorAll('[data-baseweb="tab"]');
            if (tabs.length > 1) tabs[1].click();
        }, 100);
    </script>
    """, height=0)
